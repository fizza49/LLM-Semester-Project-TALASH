"""
TALASH – CS417 Milestone 1
Preprocessing Module

Reads CV PDFs (which are ZIP archives with .txt pages),
sends text to Claude, and extracts structured data into:
  - A Python dict (for the Flask API)
  - An Excel workbook (candidates.xlsx)

Supports Claude (Anthropic), Gemini (Google), or Grok (xAI).
Set LLM_PROVIDER = "claude" / "gemini" / "grok" in config below.
"""

import os, json, re, zipfile, glob
from pathlib import Path
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── CONFIG ────────────────────────────────────────────────────────
LLM_PROVIDER   = os.environ.get("LLM_PROVIDER", "claude")   # claude | gemini | grok
ANTHROPIC_KEY  = os.environ.get("ANTHROPIC_API_KEY", "")
GEMINI_KEY     = os.environ.get("GEMINI_API_KEY", "")
GROK_KEY       = os.environ.get("GROK_API_KEY", "")

OUTPUT_EXCEL   = "candidates.xlsx"
CV_FOLDER      = "uploads"

# ── PROMPT ────────────────────────────────────────────────────────
SYSTEM_PROMPT = """You are an expert CV parser for an AI-powered HR recruitment system called TALASH.

Your task: Extract ALL structured information from a CV and return ONLY valid JSON.
No preamble, no explanation, no markdown fences — just the raw JSON object.

Return this exact schema (use null for missing fields, [] for missing arrays):

{
  "personal": {
    "name": "full name",
    "date_of_birth": "date or null",
    "nationality": "nationality or null",
    "marital_status": "status or null",
    "current_salary": "amount or null",
    "expected_salary": "amount or null",
    "present_employment": "current job or null",
    "applied_for": "position applied for or null",
    "apply_date": "date or null"
  },
  "education": [
    {
      "degree": "exact degree title",
      "level": "SSC | HSSC | Bachelor | Master | PhD | Other",
      "specialization": "field/major",
      "grade": "CGPA, percentage, or grade exactly as written",
      "passing_year": "year",
      "institution": "full institution name"
    }
  ],
  "experience": [
    {
      "role": "job title",
      "organization": "employer name",
      "location": "city/country or null",
      "start_date": "Mon-YYYY or null",
      "end_date": "Mon-YYYY or Present",
      "duration_months": estimated integer months or null
    }
  ],
  "publications": [
    {
      "title": "full paper title",
      "first_author": "name",
      "co_authors": "comma-separated names or null",
      "venue": "journal or conference name",
      "type": "Journal | Conference | Book | Other",
      "impact_factor": number or null,
      "volume": "vol number or null",
      "pages": "page range or null",
      "year": "year"
    }
  ],
  "skills": [],
  "awards": [
    {
      "type": "award/scholarship/membership",
      "detail": "description"
    }
  ],
  "references": [
    {
      "name": "full name",
      "designation": "title",
      "organization": "institution",
      "email": "email or null",
      "phone": "phone or null"
    }
  ],
  "missing_fields": ["list any important fields that are absent or unclear"]
}

Rules:
- Extract EVERYTHING visible in the CV. Do not invent data.
- For duration_months: calculate from start_date to end_date. Use current date for "Present".
- For publications: carefully distinguish Journal vs Conference.
- List all fields that seem important but are missing in missing_fields.
"""

def build_user_prompt(cv_text: str, filename: str) -> str:
    return f"Filename: {filename}\n\nCV Text:\n\n{cv_text}"


# ── TEXT EXTRACTION ───────────────────────────────────────────────
def extract_text_from_cv(pdf_path: str) -> str:
    """
    The Handler CVs are ZIP archives (.pdf extension) containing
    numbered .txt files (1.txt, 2.txt …). This function reads all
    of them in order and joins the text.

    Falls back to pdfplumber for genuine PDF files.
    """
    path = Path(pdf_path)

    # Try ZIP approach first (Handler format)
    try:
        with zipfile.ZipFile(pdf_path, 'r') as z:
            names = sorted(
                [n for n in z.namelist() if n.endswith('.txt')],
                key=lambda x: int(re.sub(r'\D', '', x) or 0)
            )
            pages = []
            for name in names:
                with z.open(name) as f:
                    pages.append(f.read().decode('utf-8', errors='ignore'))
            if pages:
                return "\n\n".join(pages)
    except Exception:
        pass

    # Fallback: genuine PDF
    try:
        import pdfplumber
        pages = []
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    pages.append(t)
        return "\n\n".join(pages)
    except Exception as e:
        print(f"  [!] Could not extract text from {path.name}: {e}")
        return ""


# ── LLM CALLERS ───────────────────────────────────────────────────
def call_claude(cv_text: str, filename: str) -> dict:
    import anthropic
    client = anthropic.Anthropic(api_key=ANTHROPIC_KEY)
    response = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=4096,
        system=SYSTEM_PROMPT,
        messages=[{"role": "user", "content": build_user_prompt(cv_text, filename)}]
    )
    return parse_llm_response(response.content[0].text)


def call_gemini(cv_text: str, filename: str) -> dict:
    """Google Gemini via REST API."""
    import urllib.request
    url = (
        "https://generativelanguage.googleapis.com/v1beta/models/"
        f"gemini-1.5-flash:generateContent?key={GEMINI_KEY}"
    )
    payload = {
        "contents": [{"parts": [{"text": SYSTEM_PROMPT + "\n\n" + build_user_prompt(cv_text, filename)}]}],
        "generationConfig": {"temperature": 0.1, "maxOutputTokens": 4096}
    }
    req = urllib.request.Request(
        url,
        data=json.dumps(payload).encode(),
        headers={"Content-Type": "application/json"},
        method="POST"
    )
    with urllib.request.urlopen(req, timeout=60) as resp:
        data = json.loads(resp.read())
    text = data["candidates"][0]["content"]["parts"][0]["text"]
    return parse_llm_response(text)


def call_grok(cv_text: str, filename: str) -> dict:
    """xAI Grok via OpenAI-compatible API."""
    import urllib.request
    url = "https://api.x.ai/v1/chat/completions"
    payload = {
        "model": "grok-3-mini",
        "messages": [
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user",   "content": build_user_prompt(cv_text, filename)}
        ],
        "max_tokens": 4096,
        "temperature": 0.1
    }
    req = urllib.request.Request(
        url,
        data=json.dumps(payload).encode(),
        headers={"Content-Type": "application/json", "Authorization": f"Bearer {GROK_KEY}"},
        method="POST"
    )
    with urllib.request.urlopen(req, timeout=60) as resp:
        data = json.loads(resp.read())
    text = data["choices"][0]["message"]["content"]
    return parse_llm_response(text)


def parse_llm_response(text: str) -> dict:
    """Strip markdown fences and parse JSON."""
    text = text.strip()
    text = re.sub(r'^```json\s*', '', text)
    text = re.sub(r'^```\s*',    '', text)
    text = re.sub(r'\s*```$',   '', text)
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        # Try to find JSON object in the response
        match = re.search(r'\{.*\}', text, re.DOTALL)
        if match:
            return json.loads(match.group())
        raise ValueError(f"Could not parse LLM response as JSON:\n{text[:300]}")


def call_llm(cv_text: str, filename: str) -> dict:
    """Route to the configured LLM provider."""
    provider = LLM_PROVIDER.lower()
    print(f"  → Calling {provider.upper()}...")
    if provider == "claude":
        return call_claude(cv_text, filename)
    elif provider == "gemini":
        return call_gemini(cv_text, filename)
    elif provider == "grok":
        return call_grok(cv_text, filename)
    else:
        raise ValueError(f"Unknown LLM_PROVIDER: {provider}. Use claude, gemini, or grok.")


# ── PROCESS ONE CV ────────────────────────────────────────────────
def process_cv(pdf_path: str) -> dict:
    """Full pipeline for a single CV file."""
    filename = Path(pdf_path).name
    print(f"\n[Processing] {filename}")

    # Step 1: Extract text
    text = extract_text_from_cv(pdf_path)
    if not text.strip():
        print(f"  [!] No text extracted from {filename}")
        return {"error": "No text extracted", "filename": filename}

    print(f"  ✓ Extracted {len(text)} characters of text")

    # Step 2: LLM parsing
    try:
        data = call_llm(text, filename)
        data["_filename"] = filename
        data["_processed_at"] = datetime.now().isoformat(timespec='seconds')
        print(f"  ✓ Parsed: {data.get('personal', {}).get('name', 'Unknown')}")
        return data
    except Exception as e:
        print(f"  [!] LLM error: {e}")
        return {"error": str(e), "filename": filename}


# ── EXCEL EXPORT ──────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="1A1A2E")
HEADER_FONT  = Font(bold=True, color="C8F04A", name="Calibri", size=10)
ACCENT_FILL  = PatternFill("solid", fgColor="16213E")
NORMAL_FONT  = Font(name="Calibri", size=10)
THIN_BORDER  = Border(
    left=Side(style='thin', color='2A2A4A'),
    right=Side(style='thin', color='2A2A4A'),
    top=Side(style='thin', color='2A2A4A'),
    bottom=Side(style='thin', color='2A2A4A'),
)

def style_header_row(ws, row_num, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER

def style_data_row(ws, row_num, num_cols, alternate=False):
    fill = PatternFill("solid", fgColor="0F0F1A") if alternate else PatternFill("solid", fgColor="13131F")
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = fill
        cell.font = Font(name="Calibri", size=10, color="D0D0E0")
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        cell.border = THIN_BORDER

def autofit_columns(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or '')))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 50)

def write_sheet(wb, sheet_name, headers, rows):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 22

    # Header
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    # Data rows
    for i, row in enumerate(rows):
        row_num = i + 2
        for col, val in enumerate(row, 1):
            ws.cell(row=row_num, column=col, value=str(val) if val is not None else "")
        style_data_row(ws, row_num, len(headers), alternate=(i % 2 == 0))
        ws.row_dimensions[row_num].height = 18

    autofit_columns(ws)
    return ws


def export_to_excel(all_candidates: list, output_path: str):
    """Write all parsed candidates into a styled multi-sheet Excel workbook."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    # ── SHEET 1: Candidates Overview ──────────────────────────────
    overview_headers = [
        "Candidate ID", "Name", "Applied For", "Apply Date",
        "Present Employment", "Current Salary", "Expected Salary",
        "Highest Degree", "Highest Institution", "Total Experience (Months)",
        "Total Publications", "Missing Fields", "File", "Processed At"
    ]
    overview_rows = []
    for cid, c in enumerate(all_candidates, 1):
        p   = c.get("personal", {}) or {}
        edu = c.get("education", []) or []
        exp = c.get("experience", []) or []
        pub = c.get("publications", []) or []
        mis = c.get("missing_fields", []) or []

        # Highest degree = last in list (usually ordered SSC→PhD)
        highest = next(
            (e for e in reversed(edu) if e.get("level") not in ("SSC", "HSSC")),
            edu[-1] if edu else {}
        )
        total_exp = sum(
            (e.get("duration_months") or 0) for e in exp
            if isinstance(e.get("duration_months"), (int, float))
        )

        overview_rows.append([
            f"C{cid:03d}",
            p.get("name"),
            p.get("applied_for"),
            p.get("apply_date"),
            p.get("present_employment"),
            p.get("current_salary"),
            p.get("expected_salary"),
            highest.get("degree"),
            highest.get("institution"),
            total_exp or None,
            len(pub),
            "; ".join(mis) if mis else None,
            c.get("_filename"),
            c.get("_processed_at"),
        ])

    write_sheet(wb, "1_Candidates_Overview", overview_headers, overview_rows)

    # ── SHEET 2: Education ────────────────────────────────────────
    edu_headers = [
        "Candidate ID", "Name", "Level", "Degree", "Specialization",
        "Grade / CGPA / %", "Passing Year", "Institution"
    ]
    edu_rows = []
    for cid, c in enumerate(all_candidates, 1):
        name = (c.get("personal") or {}).get("name", "?")
        for e in (c.get("education") or []):
            edu_rows.append([
                f"C{cid:03d}", name,
                e.get("level"), e.get("degree"), e.get("specialization"),
                e.get("grade"), e.get("passing_year"), e.get("institution")
            ])

    write_sheet(wb, "2_Education", edu_headers, edu_rows)

    # ── SHEET 3: Experience ───────────────────────────────────────
    exp_headers = [
        "Candidate ID", "Name", "Role", "Organization",
        "Location", "Start Date", "End Date", "Duration (Months)"
    ]
    exp_rows = []
    for cid, c in enumerate(all_candidates, 1):
        name = (c.get("personal") or {}).get("name", "?")
        for e in (c.get("experience") or []):
            exp_rows.append([
                f"C{cid:03d}", name,
                e.get("role"), e.get("organization"), e.get("location"),
                e.get("start_date"), e.get("end_date"), e.get("duration_months")
            ])

    write_sheet(wb, "3_Experience", exp_headers, exp_rows)

    # ── SHEET 4: Publications ─────────────────────────────────────
    pub_headers = [
        "Candidate ID", "Name", "Type", "Paper Title",
        "First Author", "Co-Authors", "Venue",
        "Impact Factor", "Volume", "Pages", "Year"
    ]
    pub_rows = []
    for cid, c in enumerate(all_candidates, 1):
        name = (c.get("personal") or {}).get("name", "?")
        for p in (c.get("publications") or []):
            pub_rows.append([
                f"C{cid:03d}", name,
                p.get("type"), p.get("title"),
                p.get("first_author"), p.get("co_authors"), p.get("venue"),
                p.get("impact_factor"), p.get("volume"),
                p.get("pages"), p.get("year")
            ])

    write_sheet(wb, "4_Publications", pub_headers, pub_rows)

    # ── SHEET 5: Awards & Memberships ────────────────────────────
    award_headers = ["Candidate ID", "Name", "Type", "Detail"]
    award_rows = []
    for cid, c in enumerate(all_candidates, 1):
        name = (c.get("personal") or {}).get("name", "?")
        for a in (c.get("awards") or []):
            award_rows.append([f"C{cid:03d}", name, a.get("type"), a.get("detail")])

    write_sheet(wb, "5_Awards", award_headers, award_rows)

    # ── SHEET 6: References ───────────────────────────────────────
    ref_headers = ["Candidate ID", "Name", "Ref Name", "Designation", "Organization", "Email", "Phone"]
    ref_rows = []
    for cid, c in enumerate(all_candidates, 1):
        name = (c.get("personal") or {}).get("name", "?")
        for r in (c.get("references") or []):
            ref_rows.append([
                f"C{cid:03d}", name,
                r.get("name"), r.get("designation"),
                r.get("organization"), r.get("email"), r.get("phone")
            ])

    write_sheet(wb, "6_References", ref_headers, ref_rows)

    # ── SHEET 7: Missing Information ──────────────────────────────
    missing_headers = ["Candidate ID", "Name", "Missing Field"]
    missing_rows = []
    for cid, c in enumerate(all_candidates, 1):
        name = (c.get("personal") or {}).get("name", "?")
        for m in (c.get("missing_fields") or []):
            missing_rows.append([f"C{cid:03d}", name, m])

    write_sheet(wb, "7_Missing_Info", missing_headers, missing_rows)

    wb.save(output_path)
    print(f"\n✓ Excel saved → {output_path}")


# ── MAIN PIPELINE ─────────────────────────────────────────────────
def run_pipeline(cv_folder: str = CV_FOLDER, output_excel: str = OUTPUT_EXCEL) -> list:
    """
    Scans cv_folder for all PDF files, processes each one,
    and writes the results to Excel.
    Returns the list of parsed candidate dicts.
    """
    # Find CVs
    pdf_files = sorted(glob.glob(os.path.join(cv_folder, "*.pdf")))
    if not pdf_files:
        print(f"[!] No PDF files found in '{cv_folder}'")
        return []

    print(f"Found {len(pdf_files)} CV(s) in '{cv_folder}'")
    print(f"LLM Provider: {LLM_PROVIDER.upper()}")
    print("=" * 60)

    results = []
    for pdf_path in pdf_files:
        candidate = process_cv(pdf_path)
        if "error" not in candidate:
            results.append(candidate)

    if results:
        export_to_excel(results, output_excel)
        print(f"\n{'='*60}")
        print(f"✓ Processed {len(results)} candidate(s) successfully.")
        print(f"✓ Output: {output_excel}")
    else:
        print("\n[!] No candidates were successfully processed.")

    return results


# ── SINGLE FILE ENTRY (for Flask API) ────────────────────────────
def process_single_cv(pdf_path: str) -> dict:
    """Called by Flask for each uploaded file. Returns parsed dict."""
    return process_cv(pdf_path)


# ── CLI ENTRY POINT ───────────────────────────────────────────────
if __name__ == "__main__":
    import sys

    # Allow overriding folder/output from command line
    folder = sys.argv[1] if len(sys.argv) > 1 else CV_FOLDER
    output = sys.argv[2] if len(sys.argv) > 2 else OUTPUT_EXCEL

    run_pipeline(folder, output)
