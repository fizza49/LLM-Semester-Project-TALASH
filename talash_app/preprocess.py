import json
import os
import re
import sys
import zipfile
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl
import pdfplumber
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SYSTEM_PROMPT = """
You are an expert CV parser for the TALASH Smart HR Recruitment System (CS417 project).
Your job is to extract structured information from university faculty/academic CVs.

You MUST return ONLY a valid JSON object. Do not include markdown, code fences, or explanations.

Return this exact JSON structure:
{
  "personal": {
    "name": "full name",
    "dob": "date of birth or null",
    "nationality": "nationality or null",
    "marital_status": "Married/Single/etc or null",
    "current_salary": "amount or null",
    "expected_salary": "amount or null",
    "present_employment": "current job or Unemployed",
    "applied_for": "exact post/position mentioned"
  },
  "education": [
    {
      "degree": "exact degree name",
      "specialization": "field/major",
      "grade_cgpa": "numeric value as string",
      "grade_type": "CGPA/Percentage",
      "passing_year": "year as string",
      "institution": "full institution name",
      "level": "SSC/HSSC/Bachelor/Master/PhD"
    }
  ],
  "experience": [
    {
      "role": "job title",
      "organization": "employer",
      "location": "city/country",
      "start_date": "Mon-YYYY or null",
      "end_date": "Mon-YYYY or Present",
      "duration_months": "estimated integer or null"
    }
  ],
  "publications": [
    {
      "title": "paper title",
      "first_author": "name",
      "co_authors": "comma-separated names",
      "venue": "journal or conference name",
      "venue_type": "Journal/Conference/Book",
      "impact_factor": "number as string or null",
      "volume": "vol number or null",
      "pages": "page range or null",
      "year": "year as string",
      "candidate_is_first_author": true
    }
  ],
  "awards": [
    {
      "type": "Scholarship/Award/Membership/etc",
      "detail": "full description"
    }
  ],
  "references": [
    {
      "name": "referee name",
      "designation": "title",
      "organization": "institution",
      "email": "email or null",
      "phone": "phone or null"
    }
  ],
  "skills": ["comma separated or list of skills"],
  "missing_fields": ["list any important fields that are absent or unclear in the CV"]
}

Rules:
- Extract only what is explicitly present in the CV text.
- Do not invent missing values.
- If grade value is numeric and <= 4.0, use grade_type = "CGPA"; if > 4, use "Percentage".
- candidate_is_first_author is true only if the candidate's name appears first in the author list.
- duration_months must be an integer when the duration is inferable, otherwise null.
- If a section does not exist, return an empty list for that section.
- missing_fields should mention important absent sections like publications, references, SSC/HSSC, skills, DOB, etc.
""".strip()


HEADER_COLOR = "1A1A2E"
HEADER_FONT = "FFFFFF"
ALT_ROW_COLOR = "F0F4FF"
THIN_BORDER = Side(style="thin", color="D0D7DE")
START_PATTERN = re.compile(r"Candidate for the Post of", re.IGNORECASE)
EDUCATION_LEVEL_RANK = {
    "ssc": 1,
    "matric": 1,
    "o level": 1,
    "hssc": 2,
    "intermediate": 2,
    "fsc": 2,
    "a level": 2,
    "bachelor": 3,
    "bs": 3,
    "bsc": 3,
    "be": 3,
    "master": 4,
    "ms": 4,
    "msc": 4,
    "mba": 4,
    "mphil": 5,
    "phd": 6,
    "doctorate": 6,
}


def call_gemini(prompt: str) -> str:
    import google.generativeai as genai

    api_key = os.environ.get("GEMINI_API_KEY", "").strip()
    if not api_key:
        raise EnvironmentError("GEMINI_API_KEY is not set.")

    model_name = os.environ.get("TALASH_GEMINI_MODEL", "gemini-3.1-flash-lite-preview").strip()
    genai.configure(api_key=api_key)

    model = genai.GenerativeModel(
        model_name=model_name,
        system_instruction=SYSTEM_PROMPT,
        generation_config={
            "temperature": 0.1,
            "response_mime_type": "application/json",
        },
    )
    response = model.generate_content(prompt)
    text = getattr(response, "text", "") or ""
    return text.strip()


def extract_text_from_pdf(pdf_path: Path) -> List[Dict[str, object]]:
    pages: List[Dict[str, object]] = []
    with pdfplumber.open(str(pdf_path)) as pdf:
        for page_no, page in enumerate(pdf.pages, start=1):
            text = (page.extract_text() or "").strip()
            pages.append({"page_no": page_no, "text": text})
    return pages


def extract_text_from_zip_pdf(zip_pdf_path: Path) -> List[Dict[str, object]]:
    pages: List[Dict[str, object]] = []
    with zipfile.ZipFile(str(zip_pdf_path), "r") as archive:
        txt_files = sorted(
            [name for name in archive.namelist() if name.lower().endswith(".txt")],
            key=lambda name: int(re.sub(r"\D", "", Path(name).stem) or 0),
        )
        for index, txt_file in enumerate(txt_files, start=1):
            with archive.open(txt_file) as file_obj:
                text = file_obj.read().decode("utf-8", errors="replace").strip()
                pages.append({"page_no": index, "text": text})
    return pages


def load_pages(path: Path) -> List[Dict[str, object]]:
    try:
        return extract_text_from_zip_pdf(path)
    except zipfile.BadZipFile:
        return extract_text_from_pdf(path)
    except Exception:
        return extract_text_from_pdf(path)


def split_candidate_sections(pages: List[Dict[str, object]], source_name: str) -> List[Dict[str, object]]:
    if not pages:
        return []

    start_indexes = [
        index for index, page in enumerate(pages)
        if START_PATTERN.search(str(page.get("text", "")))
    ]

    if not start_indexes:
        full_text = "\n\n".join(
            f"[Page {page['page_no']}]\n{page['text']}" for page in pages if page.get("text")
        ).strip()
        return [{
            "source_file": source_name,
            "candidate_key": f"{source_name}__1",
            "pages": [page["page_no"] for page in pages],
            "text": full_text,
        }] if full_text else []

    sections: List[Dict[str, object]] = []
    for section_no, start_idx in enumerate(start_indexes, start=1):
        end_idx = start_indexes[section_no] if section_no < len(start_indexes) else len(pages)
        chunk = pages[start_idx:end_idx]
        chunk_text = "\n\n".join(
            f"[Page {page['page_no']}]\n{page['text']}" for page in chunk if page.get("text")
        ).strip()
        if not chunk_text:
            continue
        sections.append(
            {
                "source_file": source_name,
                "candidate_key": f"{source_name}__{section_no}",
                "pages": [page["page_no"] for page in chunk],
                "text": chunk_text,
            }
        )
    return sections


def extract_json_block(raw: str) -> str:
    raw = raw.strip()
    raw = re.sub(r"^```(?:json)?\s*", "", raw)
    raw = re.sub(r"\s*```$", "", raw)
    if raw.startswith("{") and raw.endswith("}"):
        return raw

    start = raw.find("{")
    end = raw.rfind("}")
    if start != -1 and end != -1 and end > start:
        return raw[start:end + 1]
    return raw


def empty_record(source_file: str, page_range: str, reason: str) -> Dict[str, object]:
    return {
        "_source_file": source_file,
        "_page_range": page_range,
        "personal": {
            "name": Path(source_file).stem,
            "dob": None,
            "nationality": None,
            "marital_status": None,
            "current_salary": None,
            "expected_salary": None,
            "present_employment": None,
            "applied_for": None,
        },
        "education": [],
        "experience": [],
        "publications": [],
        "awards": [],
        "references": [],
        "skills": [],
        "missing_fields": [reason],
    }


def normalize_candidate(data: Dict[str, object], source_file: str, page_range: str) -> Dict[str, object]:
    personal = data.get("personal") if isinstance(data.get("personal"), dict) else {}

    normalized = {
        "_source_file": source_file,
        "_page_range": page_range,
        "personal": {
            "name": personal.get("name"),
            "dob": personal.get("dob"),
            "nationality": personal.get("nationality"),
            "marital_status": personal.get("marital_status"),
            "current_salary": personal.get("current_salary"),
            "expected_salary": personal.get("expected_salary"),
            "present_employment": personal.get("present_employment"),
            "applied_for": personal.get("applied_for"),
        },
        "education": data.get("education", []) if isinstance(data.get("education"), list) else [],
        "experience": data.get("experience", []) if isinstance(data.get("experience"), list) else [],
        "publications": data.get("publications", []) if isinstance(data.get("publications"), list) else [],
        "awards": data.get("awards", []) if isinstance(data.get("awards"), list) else [],
        "references": data.get("references", []) if isinstance(data.get("references"), list) else [],
        "skills": data.get("skills", []) if isinstance(data.get("skills"), list) else [],
        "missing_fields": data.get("missing_fields", []) if isinstance(data.get("missing_fields"), list) else [],
    }

    if not normalized["personal"]["name"]:
        normalized["personal"]["name"] = Path(source_file).stem

    for exp in normalized["experience"]:
        if isinstance(exp, dict):
            value = exp.get("duration_months")
            if isinstance(value, str) and value.strip().isdigit():
                exp["duration_months"] = int(value.strip())

    return normalized


def parse_candidate_text(candidate_text: str, source_file: str, pages: List[int]) -> Dict[str, object]:
    page_range = f"{pages[0]}-{pages[-1]}" if pages else "?"
    prompt = (
        f"Source file: {source_file}\n"
        f"Pages: {page_range}\n\n"
        f"Extract the candidate information from the CV text below.\n\n"
        f"{candidate_text}"
    )
    try:
        raw = call_gemini(prompt)
        parsed = json.loads(extract_json_block(raw))
        return normalize_candidate(parsed, source_file, page_range)
    except Exception as exc:
        print(f"  [!] Failed on pages {page_range}: {exc}")
        return empty_record(source_file, page_range, f"LLM extraction failed on pages {page_range}")


def collect_candidate_sections(input_path: Path) -> List[Dict[str, object]]:
    if input_path.is_file() and input_path.suffix.lower() == ".pdf":
        pages = load_pages(input_path)
        return split_candidate_sections(pages, input_path.name)

    if input_path.is_dir():
        sections: List[Dict[str, object]] = []
        for pdf_file in sorted(input_path.glob("*.pdf")):
            pages = load_pages(pdf_file)
            sections.extend(split_candidate_sections(pages, pdf_file.name))
        return sections

    raise FileNotFoundError(f"Input path not found or not supported: {input_path}")


def style_header_row(ws, row_num: int, num_cols: int) -> None:
    fill = PatternFill("solid", fgColor=HEADER_COLOR)
    font = Font(bold=True, color=HEADER_FONT, size=10)
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=THIN_BORDER, right=THIN_BORDER, top=THIN_BORDER, bottom=THIN_BORDER)


def style_data_row(ws, row_num: int, num_cols: int, alt: bool = False) -> None:
    fill = PatternFill("solid", fgColor=ALT_ROW_COLOR) if alt else None
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        if fill:
            cell.fill = fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = Border(left=THIN_BORDER, right=THIN_BORDER, top=THIN_BORDER, bottom=THIN_BORDER)


def auto_width(ws) -> None:
    for column_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)


def education_rank(entry: Dict[str, object]) -> int:
    level = str(entry.get("level", "") or "").strip().lower()
    degree = str(entry.get("degree", "") or "").strip().lower()

    for key, rank in EDUCATION_LEVEL_RANK.items():
        if key in level:
            return rank
    for key, rank in EDUCATION_LEVEL_RANK.items():
        if key in degree:
            return rank
    return 0


def get_highest_education(education: List[Dict[str, object]]) -> Dict[str, object]:
    if not education:
        return {}
    return max(education, key=education_rank)


def write_excel(all_candidates: List[Dict[str, object]], output_path: Path) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_sum = wb.create_sheet("Summary")
    sum_headers = [
        "#", "Name", "Applied For", "Highest Degree", "Specialization", "Institution",
        "CGPA/%", "Total Publications", "Years Experience", "Present Employment",
        "Source File", "Pages", "Missing Fields",
    ]
    ws_sum.append(sum_headers)
    style_header_row(ws_sum, 1, len(sum_headers))

    for index, candidate in enumerate(all_candidates, start=1):
        personal = candidate.get("personal", {})
        education = candidate.get("education", [])
        experience = candidate.get("experience", [])
        publications = candidate.get("publications", [])
        missing_fields = candidate.get("missing_fields", [])

        highest = get_highest_education(education)
        total_months = sum(
            item.get("duration_months", 0)
            for item in experience
            if isinstance(item, dict) and isinstance(item.get("duration_months"), int)
        )
        years_exp = round(total_months / 12, 1) if total_months else "—"

        ws_sum.append([
            index,
            personal.get("name", "—"),
            personal.get("applied_for", "—"),
            highest.get("degree", "—") if isinstance(highest, dict) else "—",
            highest.get("specialization", "—") if isinstance(highest, dict) else "—",
            highest.get("institution", "—") if isinstance(highest, dict) else "—",
            highest.get("grade_cgpa", "—") if isinstance(highest, dict) else "—",
            len(publications),
            years_exp,
            personal.get("present_employment", "—"),
            candidate.get("_source_file", ""),
            candidate.get("_page_range", ""),
            "; ".join(missing_fields) if missing_fields else "None",
        ])
        style_data_row(ws_sum, index + 1, len(sum_headers), alt=index % 2 == 0)
    auto_width(ws_sum)

    ws_personal = wb.create_sheet("Personal Info")
    personal_headers = [
        "#", "Source File", "Pages", "Name", "Date of Birth", "Nationality",
        "Marital Status", "Current Salary", "Expected Salary", "Present Employment", "Applied For",
    ]
    ws_personal.append(personal_headers)
    style_header_row(ws_personal, 1, len(personal_headers))
    for index, candidate in enumerate(all_candidates, start=1):
        personal = candidate.get("personal", {})
        ws_personal.append([
            index,
            candidate.get("_source_file", ""),
            candidate.get("_page_range", ""),
            personal.get("name", "—"),
            personal.get("dob", "—"),
            personal.get("nationality", "—"),
            personal.get("marital_status", "—"),
            personal.get("current_salary", "—"),
            personal.get("expected_salary", "—"),
            personal.get("present_employment", "—"),
            personal.get("applied_for", "—"),
        ])
        style_data_row(ws_personal, index + 1, len(personal_headers), alt=index % 2 == 0)
    auto_width(ws_personal)

    ws_education = wb.create_sheet("Education")
    education_headers = [
        "#", "Candidate Name", "Level", "Degree", "Specialization", "Grade / CGPA",
        "Grade Type", "Passing Year", "Institution",
    ]
    ws_education.append(education_headers)
    style_header_row(ws_education, 1, len(education_headers))
    row_num = 2
    serial = 1
    for candidate in all_candidates:
        name = candidate.get("personal", {}).get("name", candidate.get("_source_file", ""))
        for item in candidate.get("education", []):
            ws_education.append([
                serial,
                name,
                item.get("level", "—"),
                item.get("degree", "—"),
                item.get("specialization", "—"),
                item.get("grade_cgpa", "—"),
                item.get("grade_type", "—"),
                item.get("passing_year", "—"),
                item.get("institution", "—"),
            ])
            style_data_row(ws_education, row_num, len(education_headers), alt=row_num % 2 == 0)
            row_num += 1
            serial += 1
    auto_width(ws_education)

    ws_experience = wb.create_sheet("Experience")
    experience_headers = [
        "#", "Candidate Name", "Role", "Organization", "Location",
        "Start Date", "End Date", "Duration (Months)",
    ]
    ws_experience.append(experience_headers)
    style_header_row(ws_experience, 1, len(experience_headers))
    row_num = 2
    serial = 1
    for candidate in all_candidates:
        name = candidate.get("personal", {}).get("name", candidate.get("_source_file", ""))
        for item in candidate.get("experience", []):
            ws_experience.append([
                serial,
                name,
                item.get("role", "—"),
                item.get("organization", "—"),
                item.get("location", "—"),
                item.get("start_date", "—"),
                item.get("end_date", "—"),
                item.get("duration_months", "—"),
            ])
            style_data_row(ws_experience, row_num, len(experience_headers), alt=row_num % 2 == 0)
            row_num += 1
            serial += 1
    auto_width(ws_experience)

    ws_publications = wb.create_sheet("Publications")
    publication_headers = [
        "#", "Candidate Name", "Title", "First Author", "Co-Authors", "Venue",
        "Type", "Impact Factor", "Volume", "Pages", "Year", "Candidate First Author",
    ]
    ws_publications.append(publication_headers)
    style_header_row(ws_publications, 1, len(publication_headers))
    row_num = 2
    serial = 1
    for candidate in all_candidates:
        name = candidate.get("personal", {}).get("name", candidate.get("_source_file", ""))
        for item in candidate.get("publications", []):
            ws_publications.append([
                serial,
                name,
                item.get("title", "—"),
                item.get("first_author", "—"),
                item.get("co_authors", "—"),
                item.get("venue", "—"),
                item.get("venue_type", "—"),
                item.get("impact_factor", "—"),
                item.get("volume", "—"),
                item.get("pages", "—"),
                item.get("year", "—"),
                "Yes" if item.get("candidate_is_first_author") else "No",
            ])
            style_data_row(ws_publications, row_num, len(publication_headers), alt=row_num % 2 == 0)
            row_num += 1
            serial += 1
    auto_width(ws_publications)

    ws_awards = wb.create_sheet("Awards")
    award_headers = ["#", "Candidate Name", "Type", "Detail"]
    ws_awards.append(award_headers)
    style_header_row(ws_awards, 1, len(award_headers))
    row_num = 2
    serial = 1
    for candidate in all_candidates:
        name = candidate.get("personal", {}).get("name", candidate.get("_source_file", ""))
        for item in candidate.get("awards", []):
            ws_awards.append([serial, name, item.get("type", "—"), item.get("detail", "—")])
            style_data_row(ws_awards, row_num, len(award_headers), alt=row_num % 2 == 0)
            row_num += 1
            serial += 1
    auto_width(ws_awards)

    ws_references = wb.create_sheet("References")
    reference_headers = ["#", "Candidate Name", "Referee Name", "Designation", "Organization", "Email", "Phone"]
    ws_references.append(reference_headers)
    style_header_row(ws_references, 1, len(reference_headers))
    row_num = 2
    serial = 1
    for candidate in all_candidates:
        name = candidate.get("personal", {}).get("name", candidate.get("_source_file", ""))
        for item in candidate.get("references", []):
            ws_references.append([
                serial,
                name,
                item.get("name", "—"),
                item.get("designation", "—"),
                item.get("organization", "—"),
                item.get("email", "—"),
                item.get("phone", "—"),
            ])
            style_data_row(ws_references, row_num, len(reference_headers), alt=row_num % 2 == 0)
            row_num += 1
            serial += 1
    auto_width(ws_references)

    ws_skills = wb.create_sheet("Skills")
    skill_headers = ["#", "Candidate Name", "Skill"]
    ws_skills.append(skill_headers)
    style_header_row(ws_skills, 1, len(skill_headers))
    row_num = 2
    serial = 1
    for candidate in all_candidates:
        name = candidate.get("personal", {}).get("name", candidate.get("_source_file", ""))
        for skill in candidate.get("skills", []):
            ws_skills.append([serial, name, skill])
            style_data_row(ws_skills, row_num, len(skill_headers), alt=row_num % 2 == 0)
            row_num += 1
            serial += 1
    auto_width(ws_skills)

    ws_missing = wb.create_sheet("Missing Info")
    missing_headers = ["#", "Candidate Name", "Source File", "Pages", "Missing / Unclear Fields"]
    ws_missing.append(missing_headers)
    style_header_row(ws_missing, 1, len(missing_headers))
    for index, candidate in enumerate(all_candidates, start=1):
        name = candidate.get("personal", {}).get("name", candidate.get("_source_file", ""))
        missing = candidate.get("missing_fields", [])
        ws_missing.append([
            index,
            name,
            candidate.get("_source_file", ""),
            candidate.get("_page_range", ""),
            "; ".join(missing) if missing else "None",
        ])
        style_data_row(ws_missing, index + 1, len(missing_headers), alt=index % 2 == 0)
    auto_width(ws_missing)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = True

    wb.save(str(output_path))


def save_json(all_candidates: List[Dict[str, object]], output_path: Path) -> None:
    with output_path.open("w", encoding="utf-8") as file_obj:
        json.dump(all_candidates, file_obj, indent=2, ensure_ascii=False)


def run_pipeline(input_path: str, output_excel: str, output_json: Optional[str] = None) -> List[Dict[str, object]]:
    input_target = Path(input_path)
    output_excel_path = Path(output_excel)
    output_json_path = Path(output_json) if output_json else output_excel_path.with_suffix(".json")

    candidate_sections = collect_candidate_sections(input_target)
    if not candidate_sections:
        raise RuntimeError("No readable CV content found.")

    print(f"\n{'=' * 70}")
    print("TALASH Preprocessing Module - Gemini")
    print(f"Input: {input_target}")
    print(f"Candidates detected: {len(candidate_sections)}")
    print(f"{'=' * 70}\n")

    all_candidates: List[Dict[str, object]] = []
    for index, section in enumerate(candidate_sections, start=1):
        pages = section["pages"]
        page_range = f"{pages[0]}-{pages[-1]}" if pages else "?"
        print(f"[{index}/{len(candidate_sections)}] {section['source_file']} | pages {page_range}")
        candidate = parse_candidate_text(section["text"], section["source_file"], pages)
        print(f"  Name: {candidate.get('personal', {}).get('name', 'Unknown')}")
        print(f"  Education: {len(candidate.get('education', []))}")
        print(f"  Experience: {len(candidate.get('experience', []))}")
        print(f"  Publications: {len(candidate.get('publications', []))}")
        print(f"  Awards: {len(candidate.get('awards', []))}")
        print(f"  References: {len(candidate.get('references', []))}")
        print()
        all_candidates.append(candidate)

    write_excel(all_candidates, output_excel_path)
    save_json(all_candidates, output_json_path)

    print(f"Excel saved: {output_excel_path}")
    print(f"JSON saved : {output_json_path}")
    return all_candidates


def main() -> None:
    if len(sys.argv) < 3:
        print(
            "Usage:\n"
            "  python preprocess_cv_gemini.py <input_pdf_or_folder> <output_excel.xlsx> [output_json.json]\n\n"
            "Examples:\n"
            "  python preprocess_cv_gemini.py \"D:\\SEM 6\\LLM\\Project\\Handler (8).pdf\" \"talash_output.xlsx\"\n"
            "  python preprocess_cv_gemini.py \"cv_folder\" \"talash_output.xlsx\""
        )
        sys.exit(1)

    input_path = sys.argv[1]
    output_excel = sys.argv[2]
    output_json = sys.argv[3] if len(sys.argv) >= 4 else None
    run_pipeline(input_path, output_excel, output_json)


if __name__ == "__main__":
    main()
