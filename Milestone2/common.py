import json
import os
from pathlib import Path
from typing import Dict, Iterable, List, Optional

import openpyxl
from dotenv import load_dotenv
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pymongo import MongoClient

load_dotenv()


APP_ROOT = Path(__file__).resolve().parent
UPLOAD_DIR = APP_ROOT / "uploads"
OUTPUTS_DIR = APP_ROOT / "outputs"
PREPROCESS_DIR = OUTPUTS_DIR / "preprocess"
EDUCATION_DIR = OUTPUTS_DIR / "education"
PROFESSIONAL_DIR = OUTPUTS_DIR / "professional"
RESEARCH_DIR = OUTPUTS_DIR / "research"

print("MONGODB_DB =", os.getenv("MONGODB_DB"))


for path in [UPLOAD_DIR, OUTPUTS_DIR, PREPROCESS_DIR, EDUCATION_DIR, PROFESSIONAL_DIR, RESEARCH_DIR]:
    path.mkdir(parents=True, exist_ok=True)


HEADER_COLOR = "1A1A2E"
HEADER_FONT = "FFFFFF"
ALT_ROW_COLOR = "F0F4FF"
THIN_BORDER = Side(style="thin", color="D0D7DE")


def app_paths() -> Dict[str, Path]:
    return {
        "root": APP_ROOT,
        "uploads": UPLOAD_DIR,
        "outputs": OUTPUTS_DIR,
        "preprocess": PREPROCESS_DIR,
        "education": EDUCATION_DIR,
        "professional": PROFESSIONAL_DIR,
        "research": RESEARCH_DIR,
    }


def ensure_parent(path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    return path


def read_json(path: Path):
    with path.open("r", encoding="utf-8") as file_obj:
        return json.load(file_obj)


def write_json(data, path: Path) -> Path:
    ensure_parent(path)
    with path.open("w", encoding="utf-8") as file_obj:
        json.dump(data, file_obj, indent=2, ensure_ascii=False)
    return path


def mongo_enabled() -> bool:
    return bool(os.getenv("MONGODB_URI", "").strip())


def get_mongo_collection(collection_name: str):
    mongo_uri = os.getenv("MONGODB_URI", "").strip()
    mongo_db = os.getenv("MONGODB_DB", "talash").strip()
    if not mongo_uri:
        raise EnvironmentError("MONGODB_URI is not set.")
    client = MongoClient(mongo_uri)
    return client, client[mongo_db][collection_name]


def upsert_many(collection_name: str, documents: List[dict], key_fields: List[str]) -> int:
    if not documents:
        return 0

    client, collection = get_mongo_collection(collection_name)
    written = 0
    try:
        for document in documents:
            filter_doc = {key: document.get(key) for key in key_fields}
            if not all(value is not None for value in filter_doc.values()):
                collection.insert_one(document)
            else:
                collection.replace_one(filter_doc, document, upsert=True)
            written += 1
    finally:
        client.close()
    return written


def style_header_row(ws, row_num: int, num_cols: int) -> None:
    fill = PatternFill("solid", fgColor=HEADER_COLOR)
    font = Font(bold=True, color=HEADER_FONT, size=10)
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=THIN_BORDER, right=THIN_BORDER, top=THIN_BORDER, bottom=THIN_BORDER)


def style_data_row(ws, row_num: int, num_cols: int, alt: bool = False) -> None:
    fill = PatternFill("solid", fgColor=ALT_ROW_COLOR) if alt else None
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        if fill:
            cell.fill = fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = Border(left=THIN_BORDER, right=THIN_BORDER, top=THIN_BORDER, bottom=THIN_BORDER)


def auto_width(ws) -> None:
    for column_cells in ws.columns:
        max_len = 0
        letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            max_len = max(max_len, len("" if cell.value is None else str(cell.value)))
        ws.column_dimensions[letter].width = min(max(max_len + 3, 12), 45)


def write_workbook(sheets: Dict[str, List[dict]], output_path: Path) -> Path:
    ensure_parent(output_path)
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    for sheet_name, rows in sheets.items():
        ws = workbook.create_sheet(sheet_name[:31] or "Sheet1")
        if not rows:
            ws.append(["No data"])
            continue
        headers = list(rows[0].keys())
        ws.append(headers)
        style_header_row(ws, 1, len(headers))
        for idx, row in enumerate(rows, start=2):
            ws.append([row.get(header) for header in headers])
            style_data_row(ws, idx, len(headers), alt=idx % 2 == 0)
        ws.freeze_panes = "A2"
        auto_width(ws)

    workbook.save(str(output_path))
    return output_path
